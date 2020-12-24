import openpyxl
import os
from config import config


class ExcelHandler():
    def __init__(self, file_path):
        self.file_path = file_path

    def open_excel(self):
        workBook = openpyxl.load_workbook(self.file_path)
        self.workBook = workBook
        return workBook

    def get_sheet(self, name):
        workbook = self.open_excel()
        return workbook[name]

    def read_data(self, name):
        outerDict = []
        count = 0
        for row in list(self.get_sheet(name).rows):
            count += 1
            dict = {}
            if count == 1:
                titleList = [i.value for i in list(row)]
                continue
            else:
                time = 0
                for cel in row:
                    dict.setdefault(titleList[time], cel.value)
                    time += 1
            outerDict.append(dict)
        return outerDict

    def write(self, row, column, sheet, data):
        sheet = self.get_sheet(sheet)
        sheet.cell(row, column).value = data
        self.__save()
        self.__close()

    def __close(self):
        self.workBook.close()

    def __save(self):
        self.workBook.save(self.file_path)


if __name__ == '__main__':
    # 创建work_book对象
    # work_book=openpyxl.load_workbook("./cases.xlsx")
    # 获取一个sheet对象
    # sheet=work_book["login"]
    # 内容读取
    # sheet对象的cell方法（row,column）里面的value属性打印单元格的数据
    # print (sheet.cell(row=1,column=1).value)
    # 获取最大行
    # print (sheet.max_row)
    # 获取最大列
    # print(sheet.max_column)
    # 内容写入
    # sheet.cell(row=20,column=10,value="hahhaha")
    # work_book.save("./cases.xlsx")
    # print (config.EXCEL_PATH)

    excle = ExcelHandler(config.EXCEL_PATH + "\cases01.xlsx")
    print (config.EXCEL_PATH + "\cases01.xlsx")
    print(excle.read_data(name="参数管理"))
    # excle.write(1,1,"Sheet2","sdsd")
